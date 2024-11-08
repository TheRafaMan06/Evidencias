import sys
import sqlite3
from datetime import datetime
from tabulate import tabulate
import pandas as pd
import matplotlib
matplotlib.use('TkAgg') 
import matplotlib.pyplot as plt
plt.ion()  
import openpyxl
import csv
import os

def crear_tablas(conn):
    try:
        cursor = conn.cursor()
        
        # Crear tabla Unidad con restricciones (sin eliminar)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS Unidad (
                clave INTEGER PRIMARY KEY AUTOINCREMENT,
                rodada INTEGER NOT NULL CHECK (rodada IN (20, 26, 29)),
                color TEXT NOT NULL CHECK (length(color) <= 15)
            );
        """)
        
        # Crear tabla Clientes con restricciones (sin eliminar)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS Clientes (
                clave INTEGER PRIMARY KEY AUTOINCREMENT,
                apellidos TEXT NOT NULL CHECK (length(apellidos) <= 40),
                nombres TEXT NOT NULL CHECK (length(nombres) <= 40),
                telefono TEXT NOT NULL CHECK (length(telefono) = 10 AND telefono GLOB '[0-9]*')
            );
        """)
        
        # Crear tabla Prestamos con restricciones (sin eliminar)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS Prestamos (
                Folio INTEGER PRIMARY KEY AUTOINCREMENT,
                clave_unidad INTEGER NOT NULL,
                clave_cliente INTEGER NOT NULL,
                fecha_prestamo DATE NOT NULL,
                dias_prestamo INTEGER NOT NULL CHECK (dias_prestamo BETWEEN 1 AND 14),
                fecha_retorno DATE,
                FOREIGN KEY(clave_unidad) REFERENCES Unidad(clave),
                FOREIGN KEY(clave_cliente) REFERENCES Clientes(clave),
                CHECK (fecha_retorno IS NULL OR fecha_retorno >= fecha_prestamo)
            );
        """)
        
        conn.commit()
        print("Tablas verificadas exitosamente")
        return True
        
    except sqlite3.Error as e:
        print(f"Error al verificar las tablas: {e}")
        return False


def validar_fecha(fecha_str):
    try:
        fecha = datetime.strptime(fecha_str, "%m-%d-%Y")
        return True, fecha
    except ValueError:
        return False, None

def crear_conexion():
    """
    Crea una conexión a la base de datos y asegura que exista en la ruta correcta
    """
    try:
        # Usar ruta absoluta para la base de datos
        ruta_db = os.path.join(os.path.dirname(os.path.abspath(__file__)), "RentaBicicletas.db")
        conn = sqlite3.connect(ruta_db)
        
        # Habilitar las foreign keys
        conn.execute("PRAGMA foreign_keys = ON")
        
        return conn
    except sqlite3.Error as e:
        print(f"Error al crear la conexión a la base de datos: {e}")
        sys.exit(1)

def verificar_integridad_bd():
    """
    Verifica la integridad de la base de datos
    """
    try:
        with crear_conexion() as conn:
            cursor = conn.cursor()
            
            # Verificar la integridad de la base de datos
            cursor.execute("PRAGMA integrity_check")
            resultado = cursor.fetchone()
            if resultado[0] != "ok":
                print("Error: La base de datos está corrupta")
                return False
            
            # Verificar que las tablas existen y tienen la estructura correcta
            tablas_requeridas = {
                "Unidad": ["clave", "rodada", "color"],
                "Clientes": ["clave", "apellidos", "nombres", "telefono"],
                "Prestamos": ["Folio", "clave_unidad", "clave_cliente", "fecha_prestamo", 
                            "dias_prestamo", "fecha_retorno"]
            }
            
            for tabla, columnas in tablas_requeridas.items():
                cursor.execute(f"SELECT name FROM sqlite_master WHERE type='table' AND name=?", (tabla,))
                if not cursor.fetchone():
                    print(f"Error: La tabla {tabla} no existe")
                    return False
                
                cursor.execute(f"PRAGMA table_info({tabla})")
                columnas_existentes = [info[1] for info in cursor.fetchall()]
                for columna in columnas:
                    if columna not in columnas_existentes:
                        print(f"Error: La columna {columna} no existe en la tabla {tabla}")
                        return False
            
            return True
            
    except sqlite3.Error as e:
        print(f"Error al verificar la integridad de la base de datos: {e}")
        return False

def mostrar_catalogo_unidades(conn, solo_disponibles=False):
    """
    Muestra el catálogo de unidades
    Parameters:
        conn: Conexión a la base de datos
        solo_disponibles: Si es True, muestra solo las unidades disponibles para préstamo
    """
    cursor = conn.cursor()
    if solo_disponibles:
        cursor.execute("""
            SELECT u.clave, u.rodada, u.color, 
                   CASE 
                       WHEN p.clave_unidad IS NULL OR p.fecha_retorno IS NOT NULL THEN 'Disponible'
                       ELSE 'No disponible'
                   END as estado
            FROM Unidad u
            LEFT JOIN Prestamos p ON u.clave = p.clave_unidad AND p.fecha_retorno IS NULL
            ORDER BY u.clave
        """)
    else:
        cursor.execute("""
            SELECT u.clave, u.rodada, u.color,
                   CASE 
                       WHEN p.clave_unidad IS NULL OR p.fecha_retorno IS NOT NULL THEN 'Disponible'
                       ELSE 'No disponible'
                   END as estado
            FROM Unidad u
            LEFT JOIN Prestamos p ON u.clave = p.clave_unidad AND p.fecha_retorno IS NULL
            ORDER BY u.clave
        """)
    
    unidades = cursor.fetchall()
    if not unidades:
        print("\nNo hay unidades registradas" if not solo_disponibles else "\nNo hay unidades disponibles")
        return None
    
    print("\nCatálogo de Unidades:")
    print(tabulate(unidades, headers=["Clave", "Rodada", "Color", "Estado"], 
                  tablefmt="simple_grid", numalign="center", stralign="left"))
    return unidades

def mostrar_catalogo_clientes(conn):
    """
    Muestra el catálogo de clientes
    """
    cursor = conn.cursor()
    cursor.execute("""
        SELECT c.clave, c.apellidos, c.nombres, c.telefono,
               COUNT(CASE WHEN p.fecha_retorno IS NULL THEN 1 END) as prestamos_activos
        FROM Clientes c
        LEFT JOIN Prestamos p ON c.clave = p.clave_cliente
        GROUP BY c.clave
        ORDER BY c.apellidos, c.nombres
    """)
    
    clientes = cursor.fetchall()
    if not clientes:
        print("\nNo hay clientes registrados")
        return None
    
    print("\nCatálogo de Clientes:")
    print(tabulate(clientes, 
          headers=["Clave", "Apellidos", "Nombres", "Teléfono", "Préstamos Activos"],
          tablefmt="simple_grid", numalign="center", stralign="left"))
    return clientes

def solicitar_clave_unidad(conn, mensaje="Ingrese la clave de la unidad: ", solo_disponibles=False):
    """
    Solicita y valida una clave de unidad
    """
    while True:
        unidades = mostrar_catalogo_unidades(conn, solo_disponibles)
        if not unidades:
            return None
            
        clave_input = input(f"\n{mensaje} (o 'cancelar' para salir): ")
        if clave_input.lower() == 'cancelar':
            return None
            
        try:
            clave = int(clave_input)
            # Verificar que la clave existe y está disponible si es necesario
            if any(u[0] == clave for u in unidades):
                if solo_disponibles:
                    if any(u[0] == clave and u[3] == 'Disponible' for u in unidades):
                        return clave
                    else:
                        print("Error: La unidad no está disponible para préstamo")
                else:
                    return clave
            else:
                print("Error: Clave de unidad no válida")
        except ValueError:
            print("Error: Ingrese un número válido")

def solicitar_clave_cliente(conn, mensaje="Ingrese la clave del cliente: "):
    """
    Solicita y valida una clave de cliente
    """
    while True:
        clientes = mostrar_catalogo_clientes(conn)
        if not clientes:
            return None
            
        clave_input = input(f"\n{mensaje} (o 'cancelar' para salir): ")
        if clave_input.lower() == 'cancelar':
            return None
            
        try:
            clave = int(clave_input)
            if any(c[0] == clave for c in clientes):
                return clave
            else:
                print("Error: Clave de cliente no válida")
        except ValueError:
            print("Error: Ingrese un número válido")

def registrar_unidad(conn):
    try:
        print("\nPara cancelar el registro en cualquier momento, escriba 'cancelar'")
        
        # Validar rodada
        while True:
            try:
                rodada_input = input("\nIngrese la rodada (20, 26 o 29): ")
                if rodada_input.lower() == 'cancelar':
                    print("Operación cancelada")
                    return False
                    
                rodada = int(rodada_input)
                if rodada not in [20, 26, 29]:
                    print("Error: La rodada debe ser 20, 26 o 29")
                    continue
                break
            except ValueError:
                print("Error: Ingrese un número válido")

        # Validar color
        while True:
            color = input("Ingrese el color (máximo 15 caracteres): ").strip()
            if color.lower() == 'cancelar':
                print("Operación cancelada")
                return False
                
            if len(color) == 0 or len(color) > 15:
                print("Error: El color debe tener entre 1 y 15 caracteres")
                continue
                
            if not color.isalpha():
                print("Error: El color debe contener solo letras")
                continue
                
            break

        # Confirmación antes de guardar
        while True:
            confirmar = input("\n¿Desea guardar la unidad? (s/n): ").lower()
            if confirmar == 'n':
                print("Operación cancelada")
                return False
            elif confirmar == 's':
                break
            else:
                print("Por favor, ingrese 's' para confirmar o 'n' para cancelar")

        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO Unidad (rodada, color) VALUES (?, ?)",
            (rodada, color)
        )
        conn.commit()
        print(f"\nUnidad registrada exitosamente con clave: {cursor.lastrowid}")
        return True

    except sqlite3.Error as e:
        print(f"Error al registrar unidad: {e}")
        return False

    except sqlite3.Error as e:
        print(f"Error al registrar unidad: {e}")
        return False

def registrar_cliente(conn):
    try:
        print("\nPara cancelar el registro en cualquier momento, escriba 'cancelar'")
        
        # Validar apellidos
        while True:
            apellidos = input("\nIngrese los apellidos (máximo 40 caracteres): ").strip()
            if apellidos.lower() == 'cancelar':
                print("Operación cancelada")
                return False
                
            if len(apellidos) == 0 or len(apellidos) > 40:
                print("Error: Los apellidos deben tener entre 1 y 40 caracteres")
                continue
                
            if not all(c.isalpha() or c.isspace() for c in apellidos):
                print("Error: Los apellidos deben contener solo letras y espacios")
                continue
                
            break

        # Validar nombres
        while True:
            nombres = input("Ingrese los nombres (máximo 40 caracteres): ").strip()
            if nombres.lower() == 'cancelar':
                print("Operación cancelada")
                return False
                
            if len(nombres) == 0 or len(nombres) > 40:
                print("Error: Los nombres deben tener entre 1 y 40 caracteres")
                continue
                
            if not all(c.isalpha() or c.isspace() for c in nombres):
                print("Error: Los nombres deben contener solo letras y espacios")
                continue
                
            break

        # Validar teléfono
        while True:
            telefono = input("Ingrese el teléfono (10 dígitos): ").strip()
            if telefono.lower() == 'cancelar':
                print("Operación cancelada")
                return False
                
            if len(telefono) != 10:
                print("Error: El teléfono debe tener exactamente 10 dígitos")
                continue
                
            if not telefono.isdigit():
                print("Error: El teléfono debe contener solo números")
                continue
                
            break

        # Confirmación antes de guardar
        while True:
            confirmar = input("\n¿Desea guardar el cliente? (s/n): ").lower()
            if confirmar == 'n':
                print("Operación cancelada")
                return False
            elif confirmar == 's':
                break
            else:
                print("Por favor, ingrese 's' para confirmar o 'n' para cancelar")

        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO Clientes (apellidos, nombres, telefono) VALUES (?, ?, ?)",
            (apellidos, nombres, telefono)
        )
        conn.commit()
        print(f"\nCliente registrado exitosamente con clave: {cursor.lastrowid}")
        return True

    except sqlite3.Error as e:
        print(f"Error al registrar cliente: {e}")
        return False


def registrar_prestamo(conn):
    try:
        cursor = conn.cursor()
        print("\nPara cancelar el registro en cualquier momento, escriba 'cancelar'")
        
        # Obtener fecha actual del sistema
        fecha_actual = datetime.now().date()
        
        # Solicitar clave de unidad
        clave_unidad = solicitar_clave_unidad(conn, 
                                            mensaje="Seleccione la clave de la unidad a prestar: ",
                                            solo_disponibles=True)
        if clave_unidad is None:
            print("Operación cancelada")
            return False

        # Obtener datos de la unidad seleccionada
        cursor.execute("SELECT rodada, color FROM Unidad WHERE clave = ?", (clave_unidad,))
        unidad = cursor.fetchone()

        # Solicitar clave de cliente
        clave_cliente = solicitar_clave_cliente(conn, mensaje="Seleccione la clave del cliente: ")
        if clave_cliente is None:
            print("Operación cancelada")
            return False

        # Obtener datos del cliente seleccionado
        cursor.execute("SELECT nombres, apellidos FROM Clientes WHERE clave = ?", (clave_cliente,))
        cliente = cursor.fetchone()

        # Manejar la fecha del préstamo
        while True:
            usar_fecha_sistema = input("\n¿Desea usar la fecha actual del sistema? (s/n): ").lower().strip()
            if usar_fecha_sistema == 'cancelar':
                print("Operación cancelada")
                return False
                
            if usar_fecha_sistema == 's':
                fecha_prestamo = fecha_actual
                break
            elif usar_fecha_sistema == 'n':
                while True:
                    fecha_str = input("Ingrese la fecha del préstamo (mm-dd-yyyy): ")
                    if fecha_str.lower() == 'cancelar':
                        print("Operación cancelada")
                        return False
                        
                    try:
                        fecha_prestamo = datetime.strptime(fecha_str, "%m-%d-%Y").date()
                        if fecha_prestamo < fecha_actual:
                            print("Error: No se permiten fechas anteriores a la fecha actual")
                            continue
                        break
                    except ValueError:
                        print("Error: Formato de fecha inválido. Use mm-dd-yyyy")
                break
            else:
                print("Error: Ingrese 's' para sí o 'n' para no")

        # Validar días de préstamo
        while True:
            dias_input = input("\nIngrese la cantidad de días del préstamo (1-14): ")
            if dias_input.lower() == 'cancelar':
                print("Operación cancelada")
                return False
                
            try:
                dias_prestamo = int(dias_input)
                if dias_prestamo < 1 or dias_prestamo > 14:
                    print("Error: El préstamo debe ser entre 1 y 14 días")
                    continue
                break
            except ValueError:
                print("Error: Ingrese un número válido")

        # Confirmación final
        datos_prestamo = [
            ['Unidad', f"Rodada {unidad[0]}, Color {unidad[1]}"],
            ['Cliente', f"{cliente[0]} {cliente[1]}"],
            ['Fecha', fecha_prestamo.strftime('%m-%d-%Y')],
            ['Días', str(dias_prestamo)]
        ]
        
        print("\n=== RESUMEN DEL PRÉSTAMO ===")
        print(tabulate(datos_prestamo, tablefmt="simple_grid", stralign="left"))
            
        confirmar = input("\n¿Desea registrar el préstamo? (s/n): ").lower()
        if confirmar == 'n':
            print("Operación cancelada")
            return False
        elif confirmar != 's':
            print("Operación cancelada: Respuesta no válida")
            return False

        # Registrar el préstamo
        cursor.execute("""
            INSERT INTO Prestamos (
                clave_unidad, 
                clave_cliente, 
                fecha_prestamo, 
                dias_prestamo, 
                fecha_retorno
            ) VALUES (?, ?, ?, ?, ?)
        """, (clave_unidad, clave_cliente, fecha_prestamo.strftime("%m-%d-%Y"), 
              dias_prestamo, None))
        
        conn.commit()
        print(f"\nPréstamo registrado exitosamente con folio: {cursor.lastrowid}")
        return True

    except sqlite3.Error as e:
        print(f"Error en la base de datos: {e}")
        return False
    except Exception as e:
        print(f"Error: {e}")
        return False

def menu_retorno(conn):
    try:
        print("\nPara cancelar el registro en cualquier momento, escriba 'cancelar'")
        
        cursor = conn.cursor()
        cursor.execute("""
            WITH FechaEsperada AS (
                SELECT 
                    p.*,
                    date(substr(p.fecha_prestamo, 7, 4) || '-' || 
                         substr(p.fecha_prestamo, 1, 2) || '-' || 
                         substr(p.fecha_prestamo, 4, 2), 
                         '+' || p.dias_prestamo || ' days') as fecha_esperada
                FROM Prestamos p
                WHERE p.fecha_retorno IS NULL
                AND p.fecha_prestamo IS NOT NULL
            )
            SELECT 
                p.Folio,
                p.fecha_prestamo,
                u.clave as clave_unidad,
                u.rodada,
                u.color,
                c.clave as clave_cliente,
                c.nombres,
                c.apellidos,
                p.dias_prestamo,
                p.fecha_esperada
            FROM FechaEsperada p
            JOIN Unidad u ON p.clave_unidad = u.clave
            JOIN Clientes c ON p.clave_cliente = c.clave
            ORDER BY p.fecha_prestamo DESC
        """)
        prestamos_pendientes = cursor.fetchall()
        
        if not prestamos_pendientes:
            print("\nNo hay préstamos pendientes de retorno")
            return False
        
        print("\nPréstamos pendientes de retorno:")
        headers = ["Folio", "Fecha Préstamo", "Clave Unidad", "Rodada", "Color",
                  "Clave Cliente", "Nombres", "Apellidos", "Días Préstamo",
                  "Fecha Esperada"]
        print(tabulate(prestamos_pendientes, headers=headers, 
                      tablefmt="simple_grid", numalign="center", stralign="left"))

        # Validar folio
        while True:
            folio_input = input("\nIngrese el folio del préstamo a retornar: ")
            if folio_input.lower() == 'cancelar':
                print("Operación cancelada")
                return False
                
            try:
                folio = int(folio_input)
                if not any(p[0] == folio for p in prestamos_pendientes):
                    print("Error: Folio no válido")
                    continue
                break
            except ValueError:
                print("Error: Ingrese un número válido")

        # Obtener fecha del préstamo
        cursor.execute("SELECT fecha_prestamo FROM Prestamos WHERE Folio = ?", (folio,))
        fecha_prestamo = cursor.fetchone()[0]
        fecha_prestamo = datetime.strptime(fecha_prestamo, "%m-%d-%Y")

        # Validar fecha de retorno
        while True:
            fecha_input = input("\nIngrese la fecha de retorno (mm-dd-yyyy): ")
            if fecha_input.lower() == 'cancelar':
                print("Operación cancelada")
                return False
                
            try:
                fecha_retorno = datetime.strptime(fecha_input, "%m-%d-%Y")
                if fecha_retorno.date() < fecha_prestamo.date():
                    print("Error: La fecha de retorno no puede ser anterior a la fecha de préstamo")
                    continue
                break
            except ValueError:
                print("Error: Formato de fecha inválido. Use mm-dd-yyyy")

        # Confirmación final
        while True:
            datos_confirmacion = [
                ['Fecha de préstamo', fecha_prestamo.strftime('%m-%d-%Y')],
                ['Fecha de retorno', fecha_retorno.strftime('%m-%d-%Y')]
            ]
            print("\n=== CONFIRMAR RETORNO ===")
            print(tabulate(datos_confirmacion, tablefmt="simple_grid", stralign="left"))
            
            confirmar = input("\n¿Desea registrar el retorno? (s/n): ").lower()
            if confirmar == 'n':
                print("Operación cancelada")
                return False
            elif confirmar == 's':
                break
            else:
                print("Por favor, ingrese 's' para confirmar o 'n' para cancelar")

        cursor.execute("""
            UPDATE Prestamos 
            SET fecha_retorno = ? 
            WHERE Folio = ?
        """, (fecha_retorno.strftime("%m-%d-%Y"), folio))
        conn.commit()
        
        print(f"\nRetorno registrado exitosamente para el folio: {folio}")
        return True

    except sqlite3.Error as e:
        print(f"Error al registrar retorno: {e}")
        return False
    except Exception as e:
        print(f"Error: {e}")
        return False

def mostrar_ruta(*rutas):
    """Muestra la ruta de navegación actual"""
    print("\n" + " > ".join(rutas))

def inicializar_bd():
    """
    Inicializa la base de datos, creando la conexión y las tablas necesarias
    """
    try:
        conn = crear_conexion()
        if not crear_tablas(conn):
            print("Error al crear las tablas")
            return None
        return conn
    except Exception as e:
        print(f"Error al inicializar la base de datos: {e}")
        return None

def menu_principal():
    conn = None
    try:
        conn = inicializar_bd()
        if conn is None:
            print("Error: No se pudo inicializar la base de datos")
            return
            
        while True:
            mostrar_ruta("Menú Principal")
            print("\n=== MENÚ PRINCIPAL ===")
            print("1. Registro")
            print("2. Préstamo")
            print("3. Retorno")
            print("4. Informes")
            print("5. Salir")
            
            opcion = input("\nSeleccione una opción: ")
            
            if opcion == "1":
                menu_registro(conn)
            elif opcion == "2":
                menu_prestamo(conn)
            elif opcion == "3":
                menu_retorno(conn)
            elif opcion == "4":
                menu_informes(conn)
            elif opcion == "5":
                if confirmar_salida():
                    print("\n¡Hasta luego!")
                    break
            else:
                print("\nOpción no válida")
    except Exception as e:
        print(f"Error en el menú principal: {e}")
    finally:
        if conn:
            conn.close()

def menu_registro(conn):
    while True:
        mostrar_ruta("Menú Principal", "Registro")
        print("\n=== MENÚ REGISTRO ===")
        print("1. Unidad")
        print("2. Cliente")
        print("3. Volver al menú principal")
        
        opcion = input("\nSeleccione una opción: ")
        
        if opcion == "1":
            mostrar_ruta("Menú Principal", "Registro", "Unidad")
            registrar_unidad(conn)
        elif opcion == "2":
            mostrar_ruta("Menú Principal", "Registro", "Cliente")
            registrar_cliente(conn)
        elif opcion == "3":
            return
        else:
            print("\nOpción no válida")

def menu_prestamo(conn):
    mostrar_ruta("Menú Principal", "Préstamo")
    print("\n=== MENÚ PRÉSTAMO ===")
    registrar_prestamo(conn)

def menu_informes(conn):
    while True:
        mostrar_ruta("Menú Principal", "Informes")
        print("\n=== MENÚ INFORMES ===")
        print("1. Reportes")
        print("2. Análisis")
        print("3. Volver al menú principal")
        
        opcion = input("\nSeleccione una opción: ")
        
        if opcion == "1":
            menu_reportes(conn)
        elif opcion == "2":
            menu_analisis(conn)
        elif opcion == "3":
            return
        else:
            print("\nOpción no válida")


def exportar_csv(datos, headers, nombre_base):
    try:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre_archivo = f"{nombre_base}_{timestamp}.csv"
        
        with open(nombre_archivo, 'w', newline='', encoding='utf-8') as archivo:
            escritor = csv.writer(archivo)
            escritor.writerow(headers)
            escritor.writerows(datos)
        
        print(f"\nReporte exportado exitosamente como: {nombre_archivo}")
        
    except Exception as e:
        print(f"Error al exportar a CSV: {e}")

def exportar_excel(datos, headers, nombre_base):
    try:
        timestamp = datetime.now().strftime("%m%d%y_%H%M%S")
        nombre_archivo = f"{nombre_base}_{timestamp}.xlsx"
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte"
        
        # Calcular número de columnas y configurar título
        num_columnas = len(headers)
        ultima_columna_letra = openpyxl.utils.get_column_letter(num_columnas)
        
        # Agregar título
        titulo = nombre_base.replace('_', ' ').title()
        ws.merge_cells(f'A1:{ultima_columna_letra}1')
        celda_titulo = ws['A1']
        celda_titulo.value = titulo
        celda_titulo.font = openpyxl.styles.Font(bold=True, size=14)
        celda_titulo.alignment = openpyxl.styles.Alignment(horizontal='center')
        
        # Agregar fecha y hora
        fecha_hora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        ws.merge_cells(f'A2:{ultima_columna_letra}2')
        celda_fecha = ws['A2']
        celda_fecha.value = f"Generado el: {fecha_hora}"
        celda_fecha.font = openpyxl.styles.Font(italic=True)
        celda_fecha.alignment = openpyxl.styles.Alignment(horizontal='center')
        
        # Estilo base para encabezados y datos
        borde = openpyxl.styles.Border(
            left=openpyxl.styles.Side(style='thin'),
            right=openpyxl.styles.Side(style='thin'),
            top=openpyxl.styles.Side(style='thin'),
            bottom=openpyxl.styles.Side(style='thin')
        )
        
        # Agregar y dar formato a encabezados
        for col, header in enumerate(headers, 1):
            celda = ws.cell(row=4, column=col, value=header)
            celda.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            celda.fill = openpyxl.styles.PatternFill(
                start_color="366092",
                end_color="366092",
                fill_type="solid"
            )
            celda.alignment = openpyxl.styles.Alignment(horizontal='center')
            celda.border = borde
        
        # Agregar y dar formato a datos
        for row_idx, row in enumerate(datos, 5):
            for col_idx, value in enumerate(row, 1):
                celda = ws.cell(row=row_idx, column=col_idx, value=value)
                celda.alignment = openpyxl.styles.Alignment(horizontal='center')
                celda.border = borde
                if row_idx % 2 == 0:
                    celda.fill = openpyxl.styles.PatternFill(
                        start_color="F2F2F2",
                        end_color="F2F2F2",
                        fill_type="solid"
                    )
        
        # Ajustar ancho de columnas
        for col in range(1, len(headers) + 1):
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(col)
            
            for row in range(4, len(datos) + 5):
                cell = ws.cell(row=row, column=col)
                max_length = max(max_length, len(str(cell.value or '')))
            
            adjusted_width = max_length + 4
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Configurar filtros y paneles congelados
        ws.auto_filter.ref = f"A4:{ultima_columna_letra}{len(datos) + 4}"
        ws.freeze_panes = 'A5'
        
        wb.save(nombre_archivo)
        print(f"\nReporte exportado exitosamente como: {nombre_archivo}")
        
    except Exception as e:
        print(f"Error al exportar a Excel: {e}")

def menu_reportes(conn):
    while True:
        mostrar_ruta("Menú Principal", "Informes", "Reportes")
        print("\n=== MENÚ REPORTES ===")
        print("1. Clientes")
        print("2. Listado de unidades")
        print("3. Retrasos")
        print("4. Préstamos por retornar")
        print("5. Préstamos por periodo")
        print("6. Volver al menú de informes")
        
        opcion = input("\nSeleccione una opción: ")
        
        if opcion == "1":
            menu_reportes_clientes(conn)  
        elif opcion == "2":
            menu_listado_unidades(conn)
        elif opcion == "3":
            mostrar_ruta("Menú Principal", "Informes", "Reportes", "Retrasos")
            mostrar_reporte_retrasos(conn)
        elif opcion == "4":
            mostrar_ruta("Menú Principal", "Informes", "Reportes", "Préstamos por retornar")
            mostrar_prestamos_no_retornados(conn)
        elif opcion == "5":
            mostrar_ruta("Menú Principal", "Informes", "Reportes", "Préstamos por periodo")
            mostrar_prestamos_periodo(conn)
        elif opcion == "6":
            return
        else:
            print("\nOpción no válida")
            
def menu_reportes_clientes(conn):
    """Menú para los reportes de clientes"""
    while True:
        mostrar_ruta("Menú Principal", "Informes", "Reportes", "Clientes")
        print("\n=== REPORTES DE CLIENTES ===")
        print("1. Reporte completo")
        print("2. Cliente específico")
        print("3. Volver al menú de reportes")
        
        opcion = input("\nSeleccione una opción: ")
        
        if opcion == "1":
            mostrar_ruta("Menú Principal", "Informes", "Reportes", "Clientes", "Reporte completo")
            mostrar_reporte_completo_clientes(conn)
        elif opcion == "2":
            mostrar_ruta("Menú Principal", "Informes", "Reportes", "Clientes", "Cliente específico")
            mostrar_reporte_cliente_especifico(conn)
        elif opcion == "3":
            return
        else:
            print("\nOpción no válida")

def exportar_reporte(datos, headers, nombre_base):
    """
    Función para exportar reportes a CSV o Excel
    """
    while True:
        print("\n¿Desea exportar este reporte?")
        print("1. Exportar como CSV")
        print("2. Exportar como Excel")
        print("3. No exportar")
        
        opcion = input("\nSeleccione una opción: ")
        
        if opcion == "1":
            exportar_csv(datos, headers, nombre_base)
            return
        elif opcion == "2":
            exportar_excel(datos, headers, nombre_base)
            return
        elif opcion == "3":
            return
        else:
            print("Opción no válida")

def mostrar_reporte_completo_clientes(conn):
    """2.6.1. Reporte completo de clientes"""
    try:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT 
                c.clave,
                c.apellidos,
                c.nombres,
                c.telefono,
                COUNT(p.folio) as total_prestamos,
                SUM(CASE 
                    WHEN p.fecha_retorno IS NULL 
                    AND p.fecha_prestamo IS NOT NULL 
                    THEN 1 
                    ELSE 0 
                END) as prestamos_activos
            FROM Clientes c
            LEFT JOIN Prestamos p ON c.clave = p.clave_cliente
            GROUP BY c.clave, c.apellidos, c.nombres, c.telefono
            ORDER BY c.apellidos, c.nombres
        """)
        clientes = cursor.fetchall()
        
        if not clientes:
            print("\nNo hay clientes registrados en el sistema")
            return
        
        headers = ["Clave", "Apellidos", "Nombres", "Teléfono", 
                  "Total Préstamos", "Préstamos Activos"]
        print("\n=== REPORTE COMPLETO DE CLIENTES ===")
        print(tabulate(clientes, 
                      headers=headers, 
                      tablefmt="simple_grid", 
                      numalign="center",
                      stralign="left"))
        
        # Ofrecer exportación
        exportar_reporte(clientes, headers, "reporte_completo_clientes")
        
    except sqlite3.Error as e:
        print(f"Error al generar reporte completo de clientes: {e}")

def mostrar_reporte_cliente_especifico(conn):
    """2.6.2. Cliente específico"""
    try:
        # Obtener y mostrar clientes disponibles con información actualizada
        cursor = conn.cursor()
        cursor.execute("""
            WITH PrestamosActivos AS (
                SELECT 
                    clave_cliente,
                    COUNT(*) as total_prestamos,
                    SUM(CASE 
                        WHEN fecha_retorno IS NULL 
                        AND fecha_prestamo IS NOT NULL 
                        THEN 1 
                        ELSE 0 
                    END) as prestamos_activos,
                    MIN(fecha_prestamo) as primer_prestamo,
                    MAX(fecha_prestamo) as ultimo_prestamo
                FROM Prestamos
                GROUP BY clave_cliente
            )
            SELECT 
                c.clave,
                c.apellidos,
                c.nombres,
                c.telefono,
                COALESCE(p.total_prestamos, 0) as total_prestamos,
                COALESCE(p.prestamos_activos, 0) as prestamos_activos,
                p.primer_prestamo,
                p.ultimo_prestamo
            FROM Clientes c
            LEFT JOIN PrestamosActivos p ON c.clave = p.clave_cliente
            ORDER BY c.apellidos, c.nombres
        """)
        clientes = cursor.fetchall()
        
        if not clientes:
            print("\nNo hay clientes registrados en el sistema")
            return
        
        headers = ["Clave", "Apellidos", "Nombres", "Teléfono", 
                  "Total Préstamos", "Préstamos Activos"]
        print("\n=== CLIENTES DISPONIBLES ===")
        print(tabulate(clientes, headers=headers, tablefmt="simple_grid", 
                      numalign="center", stralign="left"))
        
        # Solicitar clave del cliente
        while True:
            clave_input = input("\nIngrese la clave del cliente a consultar (o 'cancelar' para salir): ")
            if clave_input.lower() == 'cancelar':
                return
                
            try:
                clave = int(clave_input)
                cliente_seleccionado = next((c for c in clientes if c[0] == clave), None)
                if cliente_seleccionado:
                    break
                else:
                    print("Error: Clave de cliente no válida")
            except ValueError:
                print("Error: Ingrese un número válido")
        
        # Mostrar información del cliente usando los datos ya obtenidos
        datos_cliente = [
            ['Clave', cliente_seleccionado[0]],
            ['Apellidos', cliente_seleccionado[1]],
            ['Nombres', cliente_seleccionado[2]],
            ['Teléfono', cliente_seleccionado[3]],
            ['Total de préstamos', cliente_seleccionado[4]],
            ['Préstamos activos', cliente_seleccionado[5]]
        ]
        
        if cliente_seleccionado[6]:  # Si hay fecha del primer préstamo
            datos_cliente.append(['Primer préstamo', cliente_seleccionado[6]])
        if cliente_seleccionado[7]:  # Si hay fecha del último préstamo
            datos_cliente.append(['Último préstamo', cliente_seleccionado[7]])
        
        print("\n=== INFORMACIÓN DEL CLIENTE ===")
        print(tabulate(datos_cliente, headers=['Campo', 'Valor'], 
                      tablefmt="simple_grid", stralign="left"))
        
        # Obtener historial de préstamos
        cursor.execute("""
            SELECT 
                p.Folio,
                p.fecha_prestamo,
                u.clave as clave_unidad,
                u.rodada,
                u.color,
                p.dias_prestamo,
                p.fecha_retorno,
                CASE 
                    WHEN p.fecha_retorno IS NULL THEN 'En préstamo'
                    WHEN julianday(p.fecha_retorno) > julianday(date(
                        substr(p.fecha_prestamo, 7, 4) || '-' || 
                        substr(p.fecha_prestamo, 1, 2) || '-' || 
                        substr(p.fecha_prestamo, 4, 2), 
                        '+' || p.dias_prestamo || ' days'
                    )) THEN 'Retornado con retraso'
                    ELSE 'Retornado a tiempo'
                END as estado
            FROM Prestamos p
            JOIN Unidad u ON p.clave_unidad = u.clave
            WHERE p.clave_cliente = ?
            ORDER BY p.fecha_prestamo DESC
        """, (clave,))
        prestamos = cursor.fetchall()
        
        if prestamos:
            headers = ["Folio", "Fecha Préstamo", "Clave Unidad", "Rodada", 
                      "Color", "Días Préstamo", "Fecha Retorno", "Estado"]
            print("\n=== HISTORIAL DE PRÉSTAMOS ===")
            print(tabulate(prestamos, headers=headers, tablefmt="simple_grid",
                         numalign="center", stralign="left"))
            
            exportar_reporte(prestamos, headers, f"historial_cliente_{clave}")
        else:
            print("\nEl cliente no tiene préstamos registrados")
        
    except sqlite3.Error as e:
        print(f"Error al generar reporte de cliente específico: {e}")

def menu_listado_unidades(conn):
    while True:
        mostrar_ruta("Menú Principal", "Informes", "Reportes", "Listado de unidades")
        print("\n=== LISTADO DE UNIDADES ===")
        print("1. Completo")
        print("2. Por rodada")
        print("3. Por color")
        print("4. Volver al menú de reportes")
        
        opcion = input("\nSeleccione una opción: ")
        
        if opcion == "1":
            mostrar_ruta("Menú Principal", "Informes", "Reportes", "Listado de unidades", "Completo")
            mostrar_listado_completo(conn)
        elif opcion == "2":
            mostrar_ruta("Menú Principal", "Informes", "Reportes", "Listado de unidades", "Por rodada")
            mostrar_listado_por_rodada(conn)
        elif opcion == "3":
            mostrar_ruta("Menú Principal", "Informes", "Reportes", "Listado de unidades", "Por color")
            mostrar_listado_por_color(conn)
        elif opcion == "4":
            return
        else:
            print("\nOpción no válida")

def analisis_duracion_prestamos(conn):
    """2.7.1. Análisis de duración de préstamos"""
    try:
        query = "SELECT dias_prestamo FROM Prestamos"
        df = pd.read_sql_query(query, conn)
        
        if df.empty:
            print("\nNo hay datos de préstamos para analizar")
            return
        
        stats = {
            "Media": df['dias_prestamo'].mean(),
            "Mediana": df['dias_prestamo'].median(),
            "Moda": df['dias_prestamo'].mode().iloc[0],
            "Mínimo": df['dias_prestamo'].min(),
            "Máximo": df['dias_prestamo'].max(),
            "Desviación Estándar": df['dias_prestamo'].std(),
            "Primer Cuartil (25%)": df['dias_prestamo'].quantile(0.25),
            "Segundo Cuartil (50%)": df['dias_prestamo'].quantile(0.50),
            "Tercer Cuartil (75%)": df['dias_prestamo'].quantile(0.75)
        }
        
        print("\n=== ANÁLISIS DE DURACIÓN DE PRÉSTAMOS ===")
        tabla_stats = [[k, round(v, 2) if pd.notnull(v) else 'N/A'] for k, v in stats.items()]
        print(tabulate(tabla_stats, headers=["Estadístico", "Valor"], 
                      tablefmt="simple_grid", numalign="right", stralign="left"))
        
    except sqlite3.Error as e:
        print(f"Error al analizar duración de préstamos: {e}")

def ranking_clientes(conn):
    """2.7.2. Ranking de clientes"""
    try:
        query = """
            SELECT 
                COUNT(p.folio) as total_prestamos,
                c.clave,
                c.nombres || ' ' || c.apellidos as nombre_completo,
                c.telefono
            FROM Clientes c
            LEFT JOIN Prestamos p ON c.clave = p.clave_cliente
            GROUP BY c.clave
            ORDER BY total_prestamos DESC
            LIMIT 10
        """
        df = pd.read_sql_query(query, conn)
        
        if df.empty:
            print("\nNo hay datos de clientes para analizar")
            return
        
        print("\n=== RANKING DE CLIENTES ===")
        print(tabulate(df.values, headers=["Total Préstamos", "Clave", "Nombre Completo", "Teléfono"]))
        
    except sqlite3.Error as e:
        print(f"Error al generar ranking de clientes: {e}")


def preferencias_por_color(conn):
    try:
        query = """
            SELECT 
                u.color,
                COUNT(p.folio) as total_prestamos
            FROM Unidad u
            LEFT JOIN Prestamos p ON u.clave = p.clave_unidad
            GROUP BY u.color
            ORDER BY total_prestamos DESC
        """
        df = pd.read_sql_query(query, conn)
        
        if df.empty:
            print("\nNo hay datos suficientes para el análisis")
            return
        
        print("\n=== PREFERENCIAS POR COLOR ===")
        print(tabulate(df.values, headers=["Color", "Total Préstamos"],
                      tablefmt="simple_grid", numalign="center", stralign="left"))
        
        plt.figure(figsize=(10, 8))
        plt.pie(df['total_prestamos'], labels=df['color'], autopct='%1.1f%%')
        plt.title('Distribución de Préstamos por Color')
        
        plt.legend(labels=[f'{c}' for c in df['color']], 
                  title="Colores",
                  loc="center left",
                  bbox_to_anchor=(1, 0, 0.5, 1))
        
        plt.draw()
        plt.pause(0.1)
        plt.show(block=False)
        input()  # Esperar input sin mensaje
        plt.close()
        
    except sqlite3.Error as e:
        print(f"Error al analizar preferencias por color: {e}")

def preferencias_por_dia(conn):
    try:
        query = """
            SELECT 
                CAST(strftime('%w', date(substr(fecha_prestamo, 7, 4) || '-' || 
                                     substr(fecha_prestamo, 1, 2) || '-' || 
                                     substr(fecha_prestamo, 4, 2))) AS INTEGER) as dia_semana,
                COUNT(*) as total_prestamos
            FROM Prestamos 
            WHERE fecha_prestamo IS NOT NULL
            GROUP BY dia_semana
            ORDER BY dia_semana
        """
        df = pd.read_sql_query(query, conn)
        
        dias_completos = pd.DataFrame({
            'dia_semana': list(range(7)),
            'total_prestamos': [0] * 7
        })
        
        if not df.empty:
            for _, row in df.iterrows():
                idx = row['dia_semana']
                dias_completos.loc[dias_completos['dia_semana'] == idx, 'total_prestamos'] = row['total_prestamos']
        
        dias = ['Domingo', 'Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado']
        dias_completos['dia'] = dias_completos['dia_semana'].apply(lambda x: dias[x])
        
        print("\n=== PREFERENCIAS POR DÍA DE LA SEMANA ===")
        print(tabulate(dias_completos[['dia', 'total_prestamos']].values, 
                      headers=["Día", "Total Préstamos"],
                      tablefmt="simple_grid", numalign="center", stralign="left"))
        
        plt.figure(figsize=(12, 6))
        bars = plt.bar(dias_completos['dia'], dias_completos['total_prestamos'])
        plt.title('Préstamos por Día de la Semana')
        plt.xticks(rotation=45)
        plt.xlabel('Día')
        plt.ylabel('Total de Préstamos')
        
        for bar in bars:
            height = bar.get_height()
            plt.text(bar.get_x() + bar.get_width()/2., height,
                    f'{int(height)}',
                    ha='center', va='bottom')
        
        plt.tight_layout()
        plt.draw()
        plt.pause(0.1)
        plt.show(block=False)
        input()  # Esperar input sin mensaje
        plt.close()
        
    except sqlite3.Error as e:
        print(f"Error al analizar preferencias por día: {e}")
        print("Query utilizada:", query)
    except Exception as e:
        print(f"Error inesperado: {e}")

def menu_preferencias_rentas(conn):
    """2.7.3. Menú de preferencias de rentas"""
    while True:
        mostrar_ruta("Menú Principal", "Informes", "Análisis", "Preferencias de rentas")
        print("\n=== PREFERENCIAS DE RENTAS ===")
        print("1. Por rodada")
        print("2. Por color")
        print("3. Por día de la semana")
        print("4. Volver al menú de análisis")
        
        opcion = input("\nSeleccione una opción: ")
        
        if opcion == "1":
            mostrar_ruta("Menú Principal", "Informes", "Análisis", "Preferencias de rentas", "Por rodada")
            preferencias_por_rodada(conn)
        elif opcion == "2":
            mostrar_ruta("Menú Principal", "Informes", "Análisis", "Preferencias de rentas", "Por color")
            preferencias_por_color(conn)
        elif opcion == "3":
            mostrar_ruta("Menú Principal", "Informes", "Análisis", "Preferencias de rentas", "Por día")
            preferencias_por_dia(conn)
        elif opcion == "4":
            return
        else:
            print("\nOpción no válida")

def menu_analisis(conn):
    while True:
        mostrar_ruta("Menú Principal", "Informes", "Análisis")
        print("\n=== MENÚ ANÁLISIS ===")
        print("1. Duración de los préstamos")
        print("2. Ranking de clientes")
        print("3. Preferencias de rentas")
        print("4. Volver al menú de informes")
        
        opcion = input("\nSeleccione una opción: ")
        
        if opcion == "1":
            mostrar_ruta("Menú Principal", "Informes", "Análisis", "Duración de los préstamos")
            analisis_duracion_prestamos(conn)
        elif opcion == "2":
            mostrar_ruta("Menú Principal", "Informes", "Análisis", "Ranking de clientes")
            ranking_clientes(conn)
        elif opcion == "3":
            menu_preferencias_rentas(conn)
        elif opcion == "4":
            return
        else:
            print("\nOpción no válida")

def preferencias_por_rodada(conn):
    try:
        query = """
            SELECT 
                u.rodada,
                COUNT(p.folio) as total_prestamos
            FROM Unidad u
            LEFT JOIN Prestamos p ON u.clave = p.clave_unidad
            GROUP BY u.rodada
            ORDER BY total_prestamos DESC
        """
        df = pd.read_sql_query(query, conn)
        
        if df.empty:
            print("\nNo hay datos suficientes para el análisis")
            return
        
        print("\n=== PREFERENCIAS POR RODADA ===")
        print(tabulate(df.values, headers=["Rodada", "Total Préstamos"],
                      tablefmt="simple_grid", numalign="center", stralign="left"))
        
        # Crear gráfica de pastel
        plt.figure(figsize=(10, 8))
        plt.pie(df['total_prestamos'], labels=df['rodada'], autopct='%1.1f%%')
        plt.title('Distribución de Préstamos por Rodada')
        
        plt.legend(labels=[f'Rodada {r}' for r in df['rodada']], 
                  title="Rodadas",
                  loc="center left",
                  bbox_to_anchor=(1, 0, 0.5, 1))
        
        plt.draw()
        plt.pause(0.1)
        plt.show(block=False)
        input()  # Esperar input sin mensaje
        plt.close()
        
    except sqlite3.Error as e:
        print(f"Error al analizar preferencias por rodada: {e}")


def confirmar_salida():
    while True:
        confirmacion = input("\n¿Está seguro que desea salir? (s/n): ").lower()
        if confirmacion == 's':
            return True
        elif confirmacion == 'n':
            return False
        else:
            print("Por favor, ingrese 's' para sí o 'n' para cancelar")

def mostrar_reporte_clientes(conn):
    """2.6.1. Reporte de clientes"""
    try:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT c.*, COUNT(p.folio) as total_prestamos
            FROM Clientes c
            LEFT JOIN Prestamos p ON c.clave = p.clave_cliente
            GROUP BY c.clave
            ORDER BY c.apellidos, c.nombres
        """)
        clientes = cursor.fetchall()
        
        if not clientes:
            print("\nNo hay clientes registrados en el sistema")
            return
        
        headers = ["Clave", "Apellidos", "Nombres", "Teléfono", "Total Préstamos"]
        print("\n=== REPORTE DE CLIENTES ===")
        print(tabulate(clientes, headers=headers))
        
        # Ofrecer exportación
        exportar_reporte(clientes, headers, "reporte_clientes")
        
    except sqlite3.Error as e:
        print(f"Error al generar reporte de clientes: {e}")


def mostrar_listado_completo(conn):
    """2.6.2. Listado completo de unidades"""
    try:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT u.*, 
                   CASE WHEN p.folio IS NULL THEN 'Disponible'
                        ELSE 'Prestada'
                   END as estado
            FROM Unidad u
            LEFT JOIN Prestamos p ON u.clave = p.clave_unidad 
                AND p.fecha_retorno IS NULL
            ORDER BY u.clave
        """)
        unidades = cursor.fetchall()
        
        if not unidades:
            print("\nNo hay unidades registradas en el sistema")
            return
        
        headers = ["Clave", "Rodada", "Color", "Estado"]
        print("\n=== LISTADO COMPLETO DE UNIDADES ===")
        print(tabulate(unidades, headers=headers))
        
        # Ofrecer exportación
        exportar_reporte(unidades, headers, "listado_unidades_completo")
        
    except sqlite3.Error as e:
        print(f"Error al generar listado de unidades: {e}")

def mostrar_listado_por_rodada(conn):
    """2.6.3. Listado de unidades por rodada"""
    try:
        while True:
            rodada = input("\nIngrese la rodada a consultar (20, 26 o 29) o 'cancelar' para salir: ")
            if rodada.lower() == 'cancelar':
                return
            
            if rodada not in ['20', '26', '29']:
                print("Error: La rodada debe ser 20, 26 o 29")
                continue
            
            cursor = conn.cursor()
            cursor.execute("""
                SELECT u.clave, u.color
                FROM Unidad u
                WHERE u.rodada = ?
                ORDER BY u.clave
            """, (rodada,))
            unidades = cursor.fetchall()
            
            if not unidades:
                print(f"\nNo hay unidades registradas con rodada {rodada}")
                continue
            
            headers = ["Clave", "Color"]
            print(f"\n=== UNIDADES CON RODADA {rodada} ===")
            print(tabulate(unidades, headers=headers))
            
            # Ofrecer exportación
            exportar_reporte(unidades, headers, f'listado_unidades_rodada_{rodada}')
            break
        
    except sqlite3.Error as e:
        print(f"Error al generar listado por rodada: {e}")


def mostrar_listado_por_color(conn):
    """2.6.4. Listado de unidades por color"""
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT DISTINCT color FROM Unidad ORDER BY color")
        colores_disponibles = cursor.fetchall()
        
        if not colores_disponibles:
            print("\nNo hay unidades registradas en el sistema")
            return
        
        # Crear un diccionario de colores para mapear números a colores
        colores_dict = {str(i+1): color[0] for i, color in enumerate(colores_disponibles)}
        
        print("\nColores disponibles:")
        for num, color in colores_dict.items():
            print(f"{num}. {color}")
        
        while True:
            seleccion = input("\nIngrese el número del color a consultar o 'cancelar' para salir: ").strip()
            if seleccion.lower() == 'cancelar':
                return
            
            if seleccion not in colores_dict:
                print(f"Error: Por favor seleccione un número entre 1 y {len(colores_dict)}")
                continue
            
            color_seleccionado = colores_dict[seleccion]
            
            cursor.execute("""
                SELECT u.clave, u.rodada
                FROM Unidad u
                WHERE u.color = ?
                ORDER BY u.clave
            """, (color_seleccionado,))
            unidades = cursor.fetchall()
            
            headers = ["Clave", "Rodada"]
            print(f"\n=== UNIDADES DE COLOR {color_seleccionado.upper()} ===")
            print(tabulate(unidades, headers=headers))
            
            # Ofrecer exportación
            exportar_reporte(unidades, headers, f'listado_unidades_color_{color_seleccionado}')
            break
        
    except sqlite3.Error as e:
        print(f"Error al generar listado por color: {e}")

def mostrar_reporte_retrasos(conn):
    """2.6.5. Reporte de retrasos"""
    try:
        fecha_actual = datetime.now().date()
        
        # Definir los headers
        headers = ["Días de Retraso", "Fecha Debida", "Clave Unidad", "Rodada", 
                  "Color", "Nombre Cliente", "Teléfono"]
        
        cursor = conn.cursor()
        cursor.execute("""
            WITH PrestamosFechas AS (
                SELECT 
                    p.*,
                    date(p.fecha_prestamo, '+' || p.dias_prestamo || ' days') as fecha_esperada
                FROM Prestamos p
            )
            SELECT 
                julianday('now') - julianday(pf.fecha_esperada) as dias_retraso,
                pf.fecha_esperada as fecha_debida,
                u.clave as clave_unidad,
                u.rodada,
                u.color,
                c.nombres || ' ' || c.apellidos as nombre_completo,
                c.telefono
            FROM PrestamosFechas pf
            JOIN Unidad u ON pf.clave_unidad = u.clave
            JOIN Clientes c ON pf.clave_cliente = c.clave
            WHERE (pf.fecha_retorno IS NULL AND date(pf.fecha_esperada) < date('now'))
               OR (pf.fecha_retorno IS NOT NULL AND julianday(pf.fecha_retorno) > julianday(pf.fecha_esperada))
            ORDER BY dias_retraso DESC
        """)
        retrasos = cursor.fetchall()
        
        if not retrasos:
            print("\nNo hay préstamos con retraso")
            return
        
        print("\n=== REPORTE DE RETRASOS ===")
        print(tabulate(retrasos, headers=headers, tablefmt="simple_grid",
                      numalign="center", stralign="left"))
        
        exportar_reporte(retrasos, headers, "reporte_retrasos")
        
    except sqlite3.Error as e:
        print(f"Error al generar reporte de retrasos: {e}")

def solicitar_periodo():
    """Función auxiliar para solicitar período de fechas"""
    while True:
        try:
            print("\nIngrese el período a consultar (o 'cancelar' para salir)")
            fecha_inicio = input("Fecha inicial (mm-dd-yyyy): ")
            if fecha_inicio.lower() == 'cancelar':
                return None, None
            
            fecha_fin = input("Fecha final (mm-dd-yyyy): ")
            if fecha_fin.lower() == 'cancelar':
                return None, None
            
            # Convertir las fechas al formato correcto
            fecha_inicio = datetime.strptime(fecha_inicio, "%m-%d-%Y").date()
            fecha_fin = datetime.strptime(fecha_fin, "%m-%d-%Y").date()
            
            if fecha_fin < fecha_inicio:
                print("Error: La fecha final no puede ser anterior a la fecha inicial")
                continue
            
            return fecha_inicio, fecha_fin
            
        except ValueError:
            print("Error: Formato de fecha inválido. Use mm-dd-yyyy")

def mostrar_prestamos_no_retornados(conn):
    try:
        fecha_inicio, fecha_fin = solicitar_periodo()
        if fecha_inicio is None:
            return
        
        headers = ["Clave Unidad", "Rodada", "Fecha Préstamo", 
                  "Nombre Cliente", "Teléfono", "Fecha Esperada"]
        
        cursor = conn.cursor()
        cursor.execute("""
            SELECT 
                u.clave as clave_unidad,
                u.rodada,
                p.fecha_prestamo,
                c.nombres || ' ' || c.apellidos as nombre_completo,
                c.telefono,
                date(substr(p.fecha_prestamo, 7, 4) || '-' || 
                     substr(p.fecha_prestamo, 1, 2) || '-' || 
                     substr(p.fecha_prestamo, 4, 2), 
                     '+' || p.dias_prestamo || ' days') as fecha_esperada
            FROM Prestamos p
            JOIN Unidad u ON p.clave_unidad = u.clave
            JOIN Clientes c ON p.clave_cliente = c.clave
            WHERE p.fecha_retorno IS NULL
              AND p.fecha_prestamo BETWEEN ? AND ?
            ORDER BY p.fecha_prestamo DESC
        """, (fecha_inicio.strftime("%m-%d-%Y"), fecha_fin.strftime("%m-%d-%Y")))
        prestamos = cursor.fetchall()
        
        if not prestamos:
            print(f"\nNo hay préstamos pendientes de retorno entre {fecha_inicio} y {fecha_fin}")
            return
        
        print(f"\n=== PRÉSTAMOS POR RETORNAR ({fecha_inicio} a {fecha_fin}) ===")
        print(tabulate(prestamos, headers=headers, tablefmt="simple_grid", 
                      numalign="center", stralign="left"))
        
        exportar_reporte(prestamos, headers, "prestamos_no_retornados")
        
    except sqlite3.Error as e:
        print(f"Error al generar reporte de préstamos por retornar: {e}")

def mostrar_prestamos_periodo(conn):
    """2.6.7. Préstamos por período"""
    try:
        fecha_inicio, fecha_fin = solicitar_periodo()
        if fecha_inicio is None:
            return
        
        headers = ["Folio", "Clave Unidad", "Rodada", "Fecha Préstamo", 
                  "Nombre Cliente", "Teléfono", "Estado"]
        
        cursor = conn.cursor()
        cursor.execute("""
            SELECT 
                p.Folio,
                u.clave as clave_unidad,
                u.rodada,
                p.fecha_prestamo,
                c.nombres || ' ' || c.apellidos as nombre_completo,
                c.telefono,
                CASE 
                    WHEN p.fecha_retorno IS NULL THEN 'En préstamo'
                    ELSE 'Retornado'
                END as estado
            FROM Prestamos p
            JOIN Unidad u ON p.clave_unidad = u.clave
            JOIN Clientes c ON p.clave_cliente = c.clave
            WHERE p.fecha_prestamo BETWEEN ? AND ?
            ORDER BY p.fecha_prestamo DESC
        """, (fecha_inicio.strftime("%m-%d-%Y"), fecha_fin.strftime("%m-%d-%Y")))
        prestamos = cursor.fetchall()
        
        if not prestamos:
            print(f"\nNo hay préstamos registrados entre {fecha_inicio} y {fecha_fin}")
            return
        
        print(f"\n=== PRÉSTAMOS DEL PERÍODO ({fecha_inicio} a {fecha_fin}) ===")
        print(tabulate(prestamos, headers=headers))
        
        exportar_reporte(prestamos, headers, "prestamos_periodo")
        
    except sqlite3.Error as e:
        print(f"Error al generar reporte de préstamos por período: {e}")

if __name__ == "__main__":
    menu_principal()
